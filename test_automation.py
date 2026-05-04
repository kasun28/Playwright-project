import asyncio
import contextlib
import openpyxl
from openpyxl.styles import PatternFill
from playwright.async_api import async_playwright
import argparse
import sys


def _normalize_output(text):
    """Normalize UI text to reduce transient whitespace/newline differences."""
    if not text:
        return ""
    return " ".join(text.replace("\n", " ").split()).strip()


def _auto_update_tc_ids(ws, default_prefix="Neg"):
    """Ensure TC IDs start with Pos/Neg and are sequential per type."""
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    header_map = {str(h).strip().lower(): idx + 1 for idx, h in enumerate(headers) if h}

    tc_col = header_map.get("tc id", 1)
    input_col = header_map.get("input", 3)

    type_col = None
    for key in ("test type", "type", "case type", "category"):
        if key in header_map:
            type_col = header_map[key]
            break

    counters = {"Pos": 1, "Neg": 1}
    changed = 0

    for row in range(2, ws.max_row + 1):
        input_val = ws.cell(row, input_col).value
        if input_val is None or str(input_val).strip() == "":
            continue

        existing_id = str(ws.cell(row, tc_col).value or "").strip().lower()
        case_type_val = ""
        if type_col is not None:
            case_type_val = str(ws.cell(row, type_col).value or "").strip().lower()

        if case_type_val.startswith("pos") or case_type_val == "positive":
            prefix = "Pos"
        elif case_type_val.startswith("neg") or case_type_val == "negative":
            prefix = "Neg"
        elif existing_id.startswith("pos"):
            prefix = "Pos"
        elif existing_id.startswith("neg"):
            prefix = "Neg"
        else:
            prefix = default_prefix

        new_id = f"{prefix}_{counters[prefix]:04d}"
        counters[prefix] += 1

        if str(ws.cell(row, tc_col).value or "").strip() != new_id:
            ws.cell(row, tc_col).value = new_id
            changed += 1

    return changed, counters["Pos"] - 1, counters["Neg"] - 1


async def _wait_for_stable_output(output_box, timeout_ms=12000, settle_ms=1200, poll_ms=200):
    """Wait until output text stops changing for settle_ms, then return it."""
    loop = asyncio.get_running_loop()
    start = loop.time()
    last_change = start
    last_text = ""

    while (loop.time() - start) * 1000 < timeout_ms:
        # Read the text from the actual paragraph node to avoid editor caret noise.
        paragraph = output_box.locator('p').first
        try:
            current_text = _normalize_output(await paragraph.text_content())
        except Exception:
            current_text = _normalize_output(await output_box.text_content())

        if current_text != last_text:
            last_text = current_text
            last_change = loop.time()
        elif (loop.time() - last_change) * 1000 >= settle_ms:
            return last_text

        await asyncio.sleep(poll_ms / 1000)

    return last_text

async def test_transliterator(excel_path, url, wait_ms=5000, type_delay_ms=80, slow_mo_ms=200, save_every=1, keep_open=False):
    """Test the Sinhala transliterator with test cases from Excel"""
    
    # Load Excel file
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Auto-update TC IDs to Pos_xxxx / Neg_xxxx format before running tests.
    changed_ids, pos_count, neg_count = _auto_update_tc_ids(ws, default_prefix="Neg")
    if changed_ids:
        wb.save(excel_path)
        print(f"Updated {changed_ids} TC IDs automatically (Pos: {pos_count}, Neg: {neg_count})")
    else:
        print(f"TC IDs already valid (Pos: {pos_count}, Neg: {neg_count})")
    
    # Find test cases (starting from row 2, skip header)
    test_cases = []
    for row in range(2, ws.max_row + 1):
        tc_id = ws.cell(row, 1).value
        if not tc_id:
            continue
        test_cases.append({
            'row': row,
            'tc_id': tc_id,
            'input': ws.cell(row, 3).value,
            'expected': ws.cell(row, 4).value
        })
    
    print(f"Found {len(test_cases)} test cases to run")
    
    # Launch browser
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        page = await browser.new_page()
        # Corrected selectors based on actual page structure
        input_box = page.locator('#singlish-input')  # Input textarea
        output_box = page.locator('.tiptap.ProseMirror')  # Output div with Sinhala text
        
        try:
            print(f"Opening {url}...")
            await page.goto(url, wait_until='domcontentloaded', timeout=30000)
            await page.wait_for_timeout(wait_ms)
            print("Page loaded successfully")
            
            passed = 0
            failed = 0
            
            # Test each case
            for idx, tc in enumerate(test_cases, 1):
                try:
                    print(f"\n[{idx}/{len(test_cases)}] Testing: {tc['tc_id']}")

                    previous_output = _normalize_output(await output_box.inner_text())

                    # Clear the input by using fill with empty string
                    await input_box.fill('')
                    await page.wait_for_timeout(300)
                    print(f"  → Cleared input")

                    # Type input slowly into the textarea
                    await input_box.type(str(tc['input']), delay=type_delay_ms)
                    print(f"  → Typed: {tc['input'][:40]}...")

                    # Wait until transliteration output stabilizes to avoid partial saves.
                    dynamic_timeout = max(wait_ms, min(12000, 2500 + len(str(tc['input'])) * 70))
                    actual_output = await _wait_for_stable_output(
                        output_box=output_box,
                        timeout_ms=dynamic_timeout,
                        settle_ms=700,
                        poll_ms=150,
                    )

                    # If output did not change yet, wait a bit more and retry once.
                    if actual_output == previous_output and str(tc['input']).strip():
                        await page.wait_for_timeout(800)
                        actual_output = await _wait_for_stable_output(
                            output_box=output_box,
                            timeout_ms=max(wait_ms, 5000),
                            settle_ms=700,
                            poll_ms=150,
                        )
                    
                    # Compare
                    status = "Pass" if actual_output == tc['expected'] else "Fail"
                    
                    # Write to Excel
                    ws.cell(tc['row'], 5).value = actual_output  # Column E: Actual output
                    ws.cell(tc['row'], 6).value = status         # Column F: Status
                    
                    if status == "Pass":
                        passed += 1
                        print(f"  ✓ PASS")
                    else:
                        failed += 1
                        print(f"  ✗ FAIL")
                        print(f"    Expected: {tc['expected'][:50]}")
                        print(f"    Got:      {actual_output[:50]}")
                    
                    # Save every N cases
                    if idx % save_every == 0:
                        wb.save(excel_path)
                        print(f"  💾 Saved progress ({idx}/{len(test_cases)})")
                    
                except Exception as e:
                    print(f"  ⚠️ Error: {str(e)}")
                    failed += 1
                    continue
            
            # Final save
            wb.save(excel_path)
            print(f"\n{'='*60}")
            print(f"RESULTS: {passed} Passed | {failed} Failed | Total: {len(test_cases)}")
            print(f"{'='*60}")
            print(f"Excel file saved: {excel_path}")
            
            if keep_open:
                print("\nBrowser kept open. Press CTRL+C to close.")
                await page.wait_for_timeout(999999)
        
        except Exception as e:
            print(f"Error: {e}")
            import traceback
            traceback.print_exc()
        
        finally:
            with contextlib.suppress(Exception):
                await browser.close()

def main():
    parser = argparse.ArgumentParser(description='Test Sinhala Transliterator')
    parser.add_argument('--excel', required=True, help='Excel file path')
    parser.add_argument('--url', required=True, help='Website URL')
    parser.add_argument('--wait-ms', type=int, default=5000, help='Wait time in ms')
    parser.add_argument('--type-delay-ms', type=int, default=80, help='Typing delay in ms')
    parser.add_argument('--slow-mo-ms', type=int, default=200, help='Slow motion in ms')
    parser.add_argument('--save-every', type=int, default=1, help='Save every N cases')
    parser.add_argument('--keep-open', action='store_true', help='Keep browser open')
    
    args = parser.parse_args()
    
    asyncio.run(test_transliterator(
        excel_path=args.excel,
        url=args.url,
        wait_ms=args.wait_ms,
        type_delay_ms=args.type_delay_ms,
        slow_mo_ms=args.slow_mo_ms,
        save_every=args.save_every,
        keep_open=args.keep_open
    ))

if __name__ == '__main__':
    main()